import openpyxl

#*Criando uma planilha no Excel
book = openpyxl.Workbook()

#*Visualizando páginas
print(book.sheetnames)

#*Criando uma página
book.create_sheet('Frutas')

#*Selecionando a página correta
frutas_page = book['Frutas']

#*Adicionando o cabeçalho e os dados
frutas_page.append(['Fruta', 'Quantidade', 'Preço'])
frutas_page.append(['Banana', 5, 'R$3,90'])
frutas_page.append(['Abacaxi', 5, 'R$4,90'])
frutas_page.append(['Pera', 5, 'R$5,90'])
frutas_page.append(['Laranja', 5, 'R$6,90'])

#*Salvando a planilha
book.save('Planilha Feira.xlsx')
